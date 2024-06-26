# 导入需要的模块
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time
import pandas as pd
from openpyxl import load_workbook
from selenium.common.exceptions import NoSuchElementException
# 创建空的数据列表用来保存数据
# 大学名称
name_all = []
# 排名
rank_all = []
# 大学性质
xingzhi_all = []
# 大学研究成果
yanjiuchenguo_all = []
# 大学学生人数
studentnumbers_all = []
#大学教员人数
teachernumbers_all = []
#国际学生人数
guojixueshengshu_all = []   Guojixueshengshu_all = []
# 综合得分
scor_all = []
# 学术声誉
xueshu_all = []
# 雇主声誉
guzhu_all = []
# 每位教员引用率
jiaoyuan_all = []
# 师生比
shishengbi_all = []
# 国际学生占比
guoji_xuesheng_all = []
# 国际教师占比
guoji_jiaoshi_all = []
#国际研究网络
guoji_yanjiuwangluo_all = []
#就业结果
jiuye_all = []
#可持续性
kechixu_all = []
def get_text():
    # 大学名称
    try:
        name_element = driver.find_element(By.XPATH,
                                           '//*[@id="block-tu-d8-content"]/div/article/div/div[1]/div/div[1]/div/div/div/div/div/div[1]/div[2]/h1')
        name = name_element.text
        name_all.append(name)
    except NoSuchElementException:
        name_all.append("无")
    # rank 排名
    try:
        rank_element = driver.find_element(By.XPATH,
                                           '//*[@id="block-tu-d8-content"]/div/article/div/div[1]/div/div[2]/div/div/div/ul/li[1]/span/b')
        rank_all.append(rank_element.text)
    except NoSuchElementException:
        rank_all.append("无")
    # xingzhi 办学性质
    try:
        xingzhi_element = driver.find_element(By.XPATH,
                                              '//*[@id="block-tu-d8-content"]/div/article/div/div[1]/div/div[2]/div/div/div/ul/li[2]/span/b')
        xingzhi_all.append(xingzhi_element.text)
    except NoSuchElementException:
        xingzhi_all.append("无")

    # 大学研究成果
    try:
        yanjiuchenguo_element = driver.find_element(By.XPATH,
                                                    '//*[@id="block-tu-d8-content"]/div/article/div/div[1]/div/div[2]/div/div/div/ul/li[3]/span/b')
        yanjiuchenguo_all.append(yanjiuchenguo_element.text)
    except NoSuchElementException:
        yanjiuchenguo_all.append("无")

    # 大学学生人数
    try:
        studentnumbers_element = driver.find_element(By.XPATH,
                                                     '//*[@id="block-tu-d8-content"]/div/article/div/div[1]/div/div[2]/div/div/div/ul/li[4]/span/b')
        studentnumbers_all.append(studentnumbers_element.text)
    except NoSuchElementException:
        studentnumbers_all.append("无")

    # 大学教员人数
    try:
        teachernumbers_element = driver.find_element(By.XPATH,
                                                     '//*[@id="block-tu-d8-content"]/div/article/div/div[1]/div/div[2]/div/div/div/ul/li[5]/span/b')
        teachernumbers_all.append(teachernumbers_element.text)
    except NoSuchElementException:
        teachernumbers_all.append("无")

    # 国际学生人数
    try:
        guojixueshengshu_element = driver.find_element(By.XPATH,
                                                       '//*[@id="block-tu-d8-content"]/div/article/div/div[1]/div/div[2]/div/div/div/ul/li[6]/span/b')
        guojixueshengshu_all.append(guojixueshengshu_element.text)
    except NoSuchElementException:
        guojixueshengshu_all.append("无")
    # 综合得分
    try:
        scor_element = driver.find_element(By.XPATH, '//*[@id="rankingsTab"]/div[3]/div[1]/div[1]')
        scor_all.append(scor_element.text)
    except NoSuchElementException:
        scor_all.append("无")

    # 学术声誉
    try:
        xueshu_element = driver.find_element(By.XPATH, '//*[@id="rankingsTab"]/div[3]/div[2]/div[1]')
        xueshu_all.append(xueshu_element.text)
    except NoSuchElementException:
        xueshu_all.append("无")

    # 雇主声誉
    try:
        guzhu_element = driver.find_element(By.XPATH, '//*[@id="rankingsTab"]/div[3]/div[3]/div[1]')
        guzhu_all.append(guzhu_element.text)
    except NoSuchElementException:
        guzhu_all.append("无")

    # 每位教员引用率
    try:
        jiaoyuan_element = driver.find_element(By.XPATH, '//*[@id="rankingsTab"]/div[3]/div[5]/div[1]')
        jiaoyuan_all.append(jiaoyuan_element.text)
    except NoSuchElementException:
        jiaoyuan_all.append("无")

    # 师生比
    try:
        shishengbi_element = driver.find_element(By.XPATH, '//*[@id="rankingsTab"]/div[3]/div[4]/div[1]')
        shishengbi_all.append(shishengbi_element.text)
    except NoSuchElementException:
        shishengbi_all.append("无")

    # 国际学生占比
    try:
        guoji_xuesheng_element = driver.find_element(By.XPATH, '//*[@id="rankingsTab"]/div[3]/div[7]/div[1]')
        guoji_xuesheng_all.append(guoji_xuesheng_element.text)
    except NoSuchElementException:
        guoji_xuesheng_all.append("无")

    # 国际教师占比
    try:
        guoji_jiaoshi_element = driver.find_element(By.XPATH, '//*[@id="rankingsTab"]/div[3]/div[6]/div[1]')
        guoji_jiaoshi_all.append(guoji_jiaoshi_element.text)
    except NoSuchElementException:
        guoji_jiaoshi_all.append("无")

    # 国际研究网络
    try:
        guoji_yanjiuwangluo_element = driver.find_element(By.XPATH, '//*[@id="rankingsTab"]/div[3]/div[8]/div[1]')driver.find_element(驱动程序)XPATH, / / * [@ id = " rankingsTab "] / div [3] / div [8] / div[1]”)
        guoji_yanjiuwangluo_all.append(guoji_yanjiuwangluo_element.text)
    except NoSuchElementException:
        guoji_yanjiuwangluo_all.append("无")

    # 就业结果
    try:
        jiuye_element = driver.find_element(By.XPATH, '//*[@id="rankingsTab"]/div[3]/div[9]/div[1]')
        jiuye_all.append(jiuye_element.text)
    except NoSuchElementException:
        jiuye_all.append("无")

    # 可持续性
    try:   试一试:
        kechixu_element = driver.find_element(By.XPATH, '//*[@id="rankingsTab"]/div[3]/div[10]/div[1]')
        kechixu_all.append(kechixu_element.text)
    except NoSuchElementException:
        kechixu_all.append("无")

    print("大学名称列表长度:", len(name_all))
    print("排名列表长度:", len(rank_all))
    print("大学性质列表长度:", len(xingzhi_all))
    print("大学研究成果列表长度:", len(yanjiuchenguo_all))
    print("大学学生人数列表长度:", len(studentnumbers_all))
    print("大学教员人数列表长度:", len(teachernumbers_all))
    print("国际学生人数列表长度:", len(guojixueshengshu_all))
    print("综合得分列表长度:", len(scor_all))
    print("学术声誉列表长度:", len(xueshu_all))
    print("雇主声誉列表长度:", len(guzhu_all))
    print("每位教员引用率列表长度:", len(jiaoyuan_all))
    print("师生比列表长度:", len(shishengbi_all))
    print("国际学生占比列表长度:", len(guoji_xuesheng_all))
    print("国际教师占比列表长度:", len(guoji_jiaoshi_all))
    print("国际研究网络列表长度:", len(guoji_yanjiuwangluo_all))
    print("就业结果列表长度:", len(jiuye_all))
    print("可持续性列表长度:", len(kechixu_all))


from selenium.webdriver.edge.service import Service从selenium.webdriver.edge.service导入服务
from selenium.webdriver.edge.options import Options从selenium.webdriver.edge.options导入选项
# 设置浏览器路径
edge_driver_path = r'F:\edgedriver_win32 (1)\msedgedriver.exe'edge_driver_path = r' f:\edgedriver_win32 (1)\msedgedriver.exe'

# 创建 EdgeOptions 对象
edge_options = Options()

# 创建 Edge 服务对象
service = Service(edge_driver_path)service = service (edge_driver_path)

# 创建 Edge 浏览器对象
driver = webdriver.Edge(service=service, options=edge_options)Driver = webdriver。边缘(服务=服务,选择= edge_options)
从selenium导入webdriver
# 获取网页从selenium.webdriver.chrome.service导入服务
file_path = r'C:\Users\15297\Desktop\2024QS世界大学排名新版.xlsx'从selenium.webdriver.common.by导入By
wb = load_workbook(filename=file_path)Wb = load_workbook(filename=file_path)wb = load_workbook(filename=file_path)   导入的时间
sheet = wb.active   活跃的   Sheet = wb.active   活跃的   导入pandas作为pd
从openpyxl导入load_workbook

for row in   在 range(2, 100):    对于范围(2,100)内的行:
    link = sheet.cell(row=row, column=3).value链接=表。细胞(行=行、列= 3)value

    print("当前链接:", link)
    driver.get   得到(link)   driver.get   得到(链接)
    get_text()
    driver.execute_script("document.documentElement.scrollTop=700")driver.execute_script (document.documentElement.scrollTop = 700)

    time.sleep(6)



get_text()






# 数据合并为数据表
data = pd.DataFrame({   Data = pd。DataFrame ({
    '名称':name_all,
    '排名':rank_all,
    '大学性质':xingzhi_all,
    '大学研究成果':yanjiuchenguo_all,
    '大学学生人数': studentnumbers_all,
    '大学教员人数':teachernumbers_all,
    '国际学生人数':guojixueshengshu_all,
    '综合得分': scor_all,
    '学术声誉': xueshu_all,
    '雇主声誉': guzhu_all,
    '每位教员引用率': jiaoyuan_all,
    '师生比': shishengbi_all,
    '国际学生占比': guoji_xuesheng_all,
    '国际教师占比': guoji_jiaoshi_all,
    '国际研究网络': guoji_yanjiuwangluo_all,
    '就业结果': jiuye_all,
    '可持续性': kechixu_all,   试一试:
name_element = driver.find_element(By.XPATH，



})除了NoSuchElementException:

# 保存为 Excel 文件，这里你可以放置你的路径
data.to_excel(r"C:\Users\15297\Desktop\尝试.xlsx", index=False)





data.head()
