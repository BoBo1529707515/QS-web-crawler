# 导入需要的模块
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time
import pandas as pd
from openpyxl import load_workbook
from selenium.common.exceptions import NoSuchElementException
# 创建空的数据列表用来保存数据
# 大学名称
name_all = []
# 排名
rank_all = []
# 大学性质
xingzhi_all = []
# 大学研究成果
yanjiuchenguo_all = []
# 大学学生人数
studentnumbers_all = []
#大学教员人数
teachernumbers_all = []
#国际学生人数
guojixueshengshu_all = []
# 综合得分
scor_all = []
# 学术声誉
xueshu_all = []
# 雇主声誉
guzhu_all = []
# 每位教员引用率
jiaoyuan_all = []
# 师生比
shishengbi_all = []
# 国际学生占比
guoji_xuesheng_all = []
# 国际教师占比
guoji_jiaoshi_all = []
#国际研究网络
guoji_yanjiuwangluo_all = []
#就业结果
jiuye_all = []
#可持续性
kechixu_all = []
#详细地址
location_all = []
#OVERVIEW
overview_all = []
#undergraduate
undergraduate_all = []
#postgraduate
postgraduate_all = []
#QS WUR Ranking By Subject
WUR_all = []
#QS Sustainability Ranking
Sustainability_all = []
#Graduate Employability Ranking
Graduate_Employability_all = []
def get_text():
    # 大学名称
    try:
        name_element = driver.find_element(By.XPATH,
                                           '//*[@id="block-tu-d8-content"]/div/article/div/div[1]/div/div[1]/div/div/div/div/div/div[1]/div[2]/h1')
        name = name_element.text
        name_all.append(name)
    except NoSuchElementException:
        name_all.append("无")
    # rank 排名
    try:
        rank_element = driver.find_element(By.XPATH,
                                           '//*[@id="block-tu-d8-content"]/div/article/div/div[1]/div/div[2]/div/div/div/ul/li[1]/span/b')
        rank_all.append(rank_element.text)
    except NoSuchElementException:
        rank_all.append("无")
    # xingzhi 办学性质
    try:
        xingzhi_element = driver.find_element(By.XPATH,
                                              '//*[@id="block-tu-d8-content"]/div/article/div/div[1]/div/div[2]/div/div/div/ul/li[2]/span/b')
        xingzhi_all.append(xingzhi_element.text)
    except NoSuchElementException:
        xingzhi_all.append("无")

    # 大学研究成果
    try:
        yanjiuchenguo_element = driver.find_element(By.XPATH,
                                                    '//*[@id="block-tu-d8-content"]/div/article/div/div[1]/div/div[2]/div/div/div/ul/li[3]/span/b')
        yanjiuchenguo_all.append(yanjiuchenguo_element.text)
    except NoSuchElementException:
        yanjiuchenguo_all.append("无")

    # 大学学生人数
    try:
        studentnumbers_element = driver.find_element(By.XPATH,
                                                     '//*[@id="block-tu-d8-content"]/div/article/div/div[1]/div/div[2]/div/div/div/ul/li[4]/span/b')
        studentnumbers_all.append(studentnumbers_element.text)
    except NoSuchElementException:
        studentnumbers_all.append("无")

    # 大学教员人数
    try:
        teachernumbers_element = driver.find_element(By.XPATH,
                                                     '//*[@id="block-tu-d8-content"]/div/article/div/div[1]/div/div[2]/div/div/div/ul/li[5]/span/b')
        teachernumbers_all.append(teachernumbers_element.text)
    except NoSuchElementException:
        teachernumbers_all.append("无")

    # 国际学生人数
    try:
        guojixueshengshu_element = driver.find_element(By.XPATH,
                                                       '//*[@id="block-tu-d8-content"]/div/article/div/div[1]/div/div[2]/div/div/div/ul/li[6]/span/b')
        guojixueshengshu_all.append(guojixueshengshu_element.text)
    except NoSuchElementException:
        guojixueshengshu_all.append("无")
    # 综合得分
    try:
        scor_element = driver.find_element(By.XPATH, '//*[@id="rankingsTab"]/div[3]/div[1]/div[1]')
        scor_all.append(scor_element.text)
    except NoSuchElementException:
        scor_all.append("无")

    # 学术声誉
    try:
        xueshu_element = driver.find_element(By.XPATH, '//*[@id="rankingsTab"]/div[3]/div[2]/div[1]')
        xueshu_all.append(xueshu_element.text)
    except NoSuchElementException:
        xueshu_all.append("无")

    # 雇主声誉
    try:
        guzhu_element = driver.find_element(By.XPATH, '//*[@id="rankingsTab"]/div[3]/div[3]/div[1]')
        guzhu_all.append(guzhu_element.text)
    except NoSuchElementException:
        guzhu_all.append("无")

    # 每位教员引用率
    try:
        jiaoyuan_element = driver.find_element(By.XPATH, '//*[@id="rankingsTab"]/div[3]/div[5]/div[1]')
        jiaoyuan_all.append(jiaoyuan_element.text)
    except NoSuchElementException:
        jiaoyuan_all.append("无")

    # 师生比
    try:
        shishengbi_element = driver.find_element(By.XPATH, '//*[@id="rankingsTab"]/div[3]/div[4]/div[1]')
        shishengbi_all.append(shishengbi_element.text)
    except NoSuchElementException:
        shishengbi_all.append("无")

    # 国际学生占比
    try:
        guoji_xuesheng_element = driver.find_element(By.XPATH, '//*[@id="rankingsTab"]/div[3]/div[7]/div[1]')
        guoji_xuesheng_all.append(guoji_xuesheng_element.text)
    except NoSuchElementException:
        guoji_xuesheng_all.append("无")

    # 国际教师占比
    try:
        guoji_jiaoshi_element = driver.find_element(By.XPATH, '//*[@id="rankingsTab"]/div[3]/div[6]/div[1]')
        guoji_jiaoshi_all.append(guoji_jiaoshi_element.text)
    except NoSuchElementException:
        guoji_jiaoshi_all.append("无")

    # 国际研究网络
    try:
        guoji_yanjiuwangluo_element = driver.find_element(By.XPATH, '//*[@id="rankingsTab"]/div[3]/div[8]/div[1]')
        guoji_yanjiuwangluo_all.append(guoji_yanjiuwangluo_element.text)
    except NoSuchElementException:
        guoji_yanjiuwangluo_all.append("无")

    # 就业结果
    try:
        jiuye_element = driver.find_element(By.XPATH, '//*[@id="rankingsTab"]/div[3]/div[9]/div[1]')
        jiuye_all.append(jiuye_element.text)
    except NoSuchElementException:
        jiuye_all.append("无")

    # 可持续性
    try:
        kechixu_element = driver.find_element(By.XPATH, '//*[@id="rankingsTab"]/div[3]/div[10]/div[1]')
        kechixu_all.append(kechixu_element.text)
    except NoSuchElementException:
        kechixu_all.append("无")
    # 详细地址
    try:
        location_element = driver.find_element(By.XPATH, '//*[@id="block-tu-d8-content"]/div/article/div/div[1]/div/div[1]/div/div/div/div/div/div[1]/div[2]/p/span')
        location_all.append( location_element.text)
    except NoSuchElementException:
        location_all.append("无")
    # OVERVIEW
    try:
        overview_element = driver.find_element(By.XPATH, '//*[@id="profile-aboutus"]/div/div/div[3]/div[2]/div[1]/div[2]')
        overview_all.append( overview_element.text)
    except NoSuchElementException:
        overview_all.append("无")
    # QS WUR Ranking By Subject

    try:
        WUR_element = driver.find_element(By.XPATH, '//*[@id="subj-tab"]/div')
        WUR_all.append( WUR_element.text)
    except NoSuchElementException:
        WUR_all.append("无")
    # QS Sustainability Ranking

    try:
        Sustainability_element = driver.find_element(By.XPATH, '//*[@id="item-3857"]/div')
        Sustainability_all.append( Sustainability_element.text)
    except NoSuchElementException:
        Sustainability_all.append("无")
    # Graduate Employability Ranking

    try:
        Graduate_Employability_element = driver.find_element(By.XPATH, '//*[@id="item-3598"]/div')
        Graduate_Employability_all.append( Graduate_Employability_element.text)
    except NoSuchElementException:
        Graduate_Employability_all.append("无")

def get_text_part2():
    # undergraduate

    try:
        undergraduate_element = driver.find_element(By.XPATH, '//*[@id="profile-aboutus"]/div/div/div[3]/div[2]/div[1]')
        undergraduate_all.append( undergraduate_element.text)
    except NoSuchElementException:
        undergraduate_all.append("无")

def get_text_part3():
    # postgraduate

    try:
        postgraduate_element = driver.find_element(By.XPATH, '//*[@id="profile-aboutus"]/div/div/div[3]/div[2]/div[1]')
        postgraduate_all.append(postgraduate_element.text)
    except NoSuchElementException:
        postgraduate_all.append("无")


##检查字长判断错误位置
    print("大学名称列表长度:", len(name_all))
    print("排名列表长度:", len(rank_all))
    print("大学性质列表长度:", len(xingzhi_all))
    print("大学研究成果列表长度:", len(yanjiuchenguo_all))
    print("大学学生人数列表长度:", len(studentnumbers_all))
    print("大学教员人数列表长度:", len(teachernumbers_all))
    print("国际学生人数列表长度:", len(guojixueshengshu_all))
    print("综合得分列表长度:", len(scor_all))
    print("学术声誉列表长度:", len(xueshu_all))
    print("雇主声誉列表长度:", len(guzhu_all))
    print("每位教员引用率列表长度:", len(jiaoyuan_all))
    print("师生比列表长度:", len(shishengbi_all))
    print("国际学生占比列表长度:", len(guoji_xuesheng_all))
    print("国际教师占比列表长度:", len(guoji_jiaoshi_all))
    print("国际研究网络列表长度:", len(guoji_yanjiuwangluo_all))
    print("就业结果列表长度:", len(jiuye_all))
    print("可持续性列表长度:", len(kechixu_all))
    print("OVERVIEW:", len(overview_all))
    print("undergraduate:", len(undergraduate_all))
    print("postgraduate:", len(postgraduate_all))
    print("QS WUR Ranking By Subject:", len(WUR_all))
    print("QS Sustainability Ranking:", len(Sustainability_all))
    print("Graduate Employability Ranking:", len(Graduate_Employability_all))

from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
# 设置浏览器路径
edge_driver_path = r'F:\edgedriver_win32 (1)\msedgedriver.exe'

# 创建 EdgeOptions 对象
edge_options = Options()

# 创建 Edge 服务对象
service = Service(edge_driver_path)

# 创建 Edge 浏览器对象
driver = webdriver.Edge(service=service, options=edge_options)

# 获取网页
file_path = r'C:\Users\15297\Desktop\2024QS世界大学排名新版.xlsx'
wb = load_workbook(filename=file_path)
sheet = wb.active

# 爬取每个链接对应页面的信息
for row in range(2, 100):  # 读取前n行
    link = sheet.cell(row=row, column=3).value

    print("当前链接:", link)
    driver.get(link)
    ## 控制鼠标点击更多控制视窗挪到写有大学地址的元素之下
    element = driver.find_element(By.XPATH, '//*[@id="block-tu-d8-content"]/div/article/div/div[1]/div/div[1]/div/div/div/div/div/div[1]/div[1]/div/img')
    driver.execute_script("arguments[0].scrollIntoView();", element)



    ## 控制鼠标点击更多
    try:
        element = driver.find_element(By.XPATH,'(//*[@id="profile-aboutus"]/div/div/div[3]/div[2]/div[1]/div[1]/p[last()]/span/a | //*[@id="profile-aboutus"]/div/div/div[2]/div[4]/div/div/ul/span/a[contains(text(), "Read more")])[1]')
        driver.execute_script("arguments[0].click();", element)
    except NoSuchElementException:
        print("找不到指定的XPath元素，跳过相应操作。")
    time.sleep(1)
    ##获得第一部分的数据
    get_text()
    time.sleep(1)
    element = driver.find_element(By.XPATH,'//*[@id="block-tu-d8-content"]/div/article/div/div[1]/div/div[1]/div/div/div/div/div/div[1]/div[1]/div/img')
    driver.execute_script("arguments[0].scrollIntoView();", element)
    ##换部分

    try:
        # 点击undergraduate的按钮
        element = driver.find_element(By.XPATH,'//*[@id="block-tu-d8-content"]/div/article/div/div[1]/div/div[3]/div/div/ul/li/a[contains(text(), "UNDERGRADUATE")]')
        driver.execute_script("arguments[0].click();", element)

        # 控制鼠标点击更多控制视窗挪到写有大学地址的元素之下
        element = driver.find_element(By.XPATH,'//*[@id="block-tu-d8-content"]/div/article/div/div[1]/div/div[1]/div/div/div/div/div/div[1]/div[1]/div/img')
        driver.execute_script("arguments[0].scrollIntoView();", element)

        # 控制鼠标点击更多
        try:
             driver.find_element(By.XPATH,'//*[@id="profile-aboutus"]/div/div/div[3]/div[2]/div[1]/div[1]/p[last()]/span/a').click()
        except NoSuchElementException:
            print("找不到指定的XPath元素，跳过相应操作。")
    except NoSuchElementException:
        print("找不到指定的XPath元素，跳过相应操作。")

    # 获得第二部分的数据
    get_text_part2()
    # 通过设置sleep时间来控制爬虫的速度
    time.sleep(1)
    try:
        ##换部分
        element = driver.find_element(By.XPATH,'//*[@id="block-tu-d8-content"]/div/article/div/div[1]/div/div[1]/div/div/div/div/div/div[1]/div[1]/div/img')
        driver.execute_script("arguments[0].scrollIntoView();", element)
        ##点击postgurate的按钮
        driver.find_element(By.XPATH, '//*[@id="block-tu-d8-content"]/div/article/div/div[1]/div/div[3]/div/div/ul/li/a[contains(text(), "POSTGRADUATE")]').click()

        ## 控制鼠标点击更多控制视窗挪到about之下
        element = driver.find_element(By.XPATH,'//*[@id="profile-aboutus"]/div/div/div[last()]/div[1]/div/h2')
        driver.execute_script("arguments[0].scrollIntoView();", element)
        ## 控制鼠标点击更多
        try:
            time.sleep(1)
            element = driver.find_element(By.XPATH,'(//*[@id="profile-aboutus"]/div/div/div[3]/div[2]/div[1]/div[1]/p[last()]/span/a | //*[@id="profile-aboutus"]/div/div/div[2]/div[4]/div/div/ul/span/a[contains(text(), "Read more")])[1]')
            driver.execute_script("arguments[0].click();", element)
        except NoSuchElementException:
            print("找不到指定的XPath元素，跳过相应操作。")
    except NoSuchElementException:
        print("找不到指定的XPath元素，跳过相应操作。")
    ##获得第三部分的数据
    get_text_part3()

get_text()
get_text_part2()
get_text_part3()




# 数据合并为数据表
data = pd.DataFrame({
    '名称':name_all,
    '排名':rank_all,
    '大学性质':xingzhi_all,
    '大学研究成果':yanjiuchenguo_all,
    '大学学生人数': studentnumbers_all,
    '大学教员人数':teachernumbers_all,
    '国际学生人数':guojixueshengshu_all,
    '综合得分': scor_all,
    '学术声誉': xueshu_all,
    '雇主声誉': guzhu_all,
    '每位教员引用率': jiaoyuan_all,
    '师生比': shishengbi_all,
    '国际学生占比': guoji_xuesheng_all,
    '国际教师占比': guoji_jiaoshi_all,
    '国际研究网络': guoji_yanjiuwangluo_all,
    '就业结果': jiuye_all,
    '可持续性': kechixu_all,
    '详细地址': location_all,
    'QS WUR Ranking By Subject': WUR_all ,
    'QS Sustainability Ranking': Sustainability_all ,
    'Graduate Employability Ranking': Graduate_Employability_all,
    'overview': overview_all,
    'undergraduate':undergraduate_all,
    'postgraduate':postgraduate_all,
})

# 保存为 Excel 文件
data.to_excel(r"C:\Users\15297\Desktop\尝试.xlsx", index=False)





data.head()